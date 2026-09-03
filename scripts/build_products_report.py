import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

root = Path(__file__).resolve().parents[1]
raw = json.loads((root / "products-export.json").read_text(encoding="utf-8"))
products = raw.get("result", {}).get("data", {}).get("json", [])
if not isinstance(products, list):
    raise SystemExit("Unexpected products response")

fields = [
    ("id", "المعرف"), ("barcode", "الباركود"), ("name", "اسم المنتج"),
    ("size", "المقاس"), ("color", "اللون"), ("weightGrams", "الوزن (جم)"),
    ("yarnDetails", "تفاصيل الخيوط"), ("imageUrl", "الصورة"),
    ("attachments", "المرفقات"), ("isActive", "نشط"),
    ("createdAt", "تاريخ الإضافة"), ("updatedAt", "آخر تحديث"),
]
required = {"name": "اسم المنتج", "size": "المقاس", "color": "اللون", "weightGrams": "الوزن", "yarnDetails": "تفاصيل الخيوط"}

def text(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)

def missing(p):
    out = []
    for key, label in required.items():
        v = p.get(key)
        if v is None or v == "" or v == 0 or (key == "yarnDetails" and (not isinstance(v, dict) or not v)):
            out.append(label)
    return "، ".join(out) if out else "لا يوجد"

for p in products:
    p["missingFields"] = missing(p)

wb = Workbook()
ws = wb.active
ws.title = "دليل المنتجات"
headers = [label for _, label in fields] + ["البيانات الناقصة"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A7EA4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
for p in products:
    ws.append([text(p.get(key)) for key, _ in fields] + [p["missingFields"]])
for column in ws.columns:
    letter = column[0].column_letter
    ws.column_dimensions[letter].width = min(max(max(len(text(c.value)) for c in column) + 2, 12), 45)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

summary = wb.create_sheet("ملخص النواقص")
summary.append(["البند", "القيمة"])
summary.append(["إجمالي المنتجات", len(products)])
summary.append(["منتجات مكتملة الحقول الأساسية", sum(1 for p in products if p["missingFields"] == "لا يوجد")])
summary.append(["منتجات تحتاج مراجعة", sum(1 for p in products if p["missingFields"] != "لا يوجد")])
summary.append([])
summary.append(["الباركود", "اسم المنتج", "المقاس", "اللون", "البيانات الناقصة"])
for cell in summary[1] + summary[6]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A7EA4")
for p in products:
    if p["missingFields"] != "لا يوجد":
        summary.append([text(p.get("barcode")), text(p.get("name")), text(p.get("size")), text(p.get("color")), p["missingFields"]])
for col in "ABCDE":
    summary.column_dimensions[col].width = 28
summary.freeze_panes = "A7"

xlsx = root / "تقرير-دليل-المنتجات-الكامل.xlsx"
wb.save(xlsx)

md = ["# تقرير دليل المنتجات الكامل", "", f"إجمالي المنتجات: **{len(products)}**", f"المنتجات المكتملة: **{sum(1 for p in products if p['missingFields'] == 'لا يوجد')}**", f"المنتجات التي تحتاج مراجعة: **{sum(1 for p in products if p['missingFields'] != 'لا يوجد')}**", "", "| الباركود | اسم المنتج | المقاس | اللون | الوزن | الخيوط | الصورة | المرفقات | البيانات الناقصة |", "|---|---|---|---|---:|---|---|---|---|"]
for p in products:
    md.append("| " + " | ".join([
        text(p.get("barcode")), text(p.get("name")), text(p.get("size")), text(p.get("color")),
        text(p.get("weightGrams")), text(p.get("yarnDetails")), "نعم" if p.get("imageUrl") else "لا",
        "نعم" if p.get("attachments") else "لا", p["missingFields"]
    ]) + " |")
(root / "تقرير-دليل-المنتجات-الكامل.md").write_text("\n".join(md) + "\n", encoding="utf-8")
print(f"products={len(products)}")
print(f"xlsx={xlsx}")
print(f"markdown={root / 'تقرير-دليل-المنتجات-الكامل.md'}")
